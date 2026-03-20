from datetime import datetime
from pathlib import Path

import openpyxl
from django.conf import settings
from django.db import transaction

from pharmacy.models import Medication, Patient


HEADERS = [
    "patient_id",
    "name",
    "age",
    "phone",
    "language",
    "drug_name",
    "dosage",
    "Indication",
    "Direction",
    "refill_due",
    "Indication_Read",
    "Direction_Read",
    "Prescriber",
    "notes",
]


def excel_file_path():
    return Path(settings.EXCEL_SYNC_FILE)


@transaction.atomic
def import_patients_from_excel(file_path=None):
    path = Path(file_path or excel_file_path())
    if not path.exists():
        return {"created": 0, "updated": 0, "rows": 0, "patients": 0}

    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in worksheet[1]]
    created = 0
    updated = 0
    deleted = 0
    removed_medications = 0
    rows = 0
    touched_patients = set()
    imported_medications = {}

    def get_value(row, name):
        try:
            item = row[headers.index(name)]
        except ValueError:
            return None
        return item.value if hasattr(item, "value") else item

    for row in worksheet.iter_rows(min_row=2, values_only=False):
        patient_id = str(get_value(row, "patient_id") or "").strip()
        if not patient_id:
            continue

        patient, was_created = Patient.objects.get_or_create(
            patient_id=patient_id,
            defaults={
                "name": str(get_value(row, "name") or "").strip(),
                "age": int(get_value(row, "age")) if get_value(row, "age") else None,
                "phone": str(get_value(row, "phone") or "").strip(),
                "language": str(get_value(row, "language") or "English").strip(),
            },
        )
        if was_created:
            created += 1
        else:
            patient.name = str(get_value(row, "name") or patient.name).strip()
            patient.age = int(get_value(row, "age")) if get_value(row, "age") else patient.age
            patient.phone = str(get_value(row, "phone") or patient.phone).strip()
            patient.language = str(get_value(row, "language") or patient.language).strip()
            patient.save()
            updated += 1

        refill_due = get_value(row, "refill_due")
        if isinstance(refill_due, datetime):
            refill_due = refill_due.date()

        drug_name = str(get_value(row, "drug_name") or "").strip()
        dosage = str(get_value(row, "dosage") or "").strip()
        if drug_name:
            Medication.objects.update_or_create(
                patient=patient,
                drug_name=drug_name,
                dosage=dosage,
                defaults={
                    "indication": str(get_value(row, "Indication") or "").strip(),
                    "direction": str(get_value(row, "Direction") or "").strip(),
                    "refill_due": refill_due,
                    "indication_read": bool(get_value(row, "Indication_Read") or False),
                    "direction_read": bool(get_value(row, "Direction_Read") or False),
                    "prescriber": str(get_value(row, "Prescriber") or "").strip(),
                    "notes": str(get_value(row, "notes") or "").strip(),
                },
            )
            imported_medications.setdefault(patient.patient_id, set()).add((drug_name.casefold(), dosage.casefold()))
        rows += 1
        touched_patients.add(patient_id)

    for patient_id in touched_patients:
        patient = Patient.objects.prefetch_related("medications").get(patient_id=patient_id)
        valid_keys = imported_medications.get(patient_id, set())
        for medication in patient.medications.exclude(status="yellow"):
            med_key = (medication.drug_name.casefold(), medication.dosage.casefold())
            if med_key not in valid_keys:
                medication.delete()
                removed_medications += 1

    stale_patients = Patient.objects.exclude(patient_id__in=touched_patients)
    for patient in stale_patients:
        if patient.call_sessions.exists():
            continue
        patient.delete()
        deleted += 1

    return {
        "created": created,
        "updated": updated,
        "deleted": deleted,
        "removed_medications": removed_medications,
        "rows": rows,
        "patients": len(touched_patients),
    }


def export_patients_to_excel(file_path=None):
    path = Path(file_path or excel_file_path())
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Patients"
    worksheet.append(HEADERS)

    for patient in Patient.objects.prefetch_related("medications").order_by("patient_id"):
        medications = list(patient.medications.all()) or [None]
        for medication in medications:
            worksheet.append(
                [
                    patient.patient_id,
                    patient.name,
                    patient.age,
                    patient.phone,
                    patient.language,
                    getattr(medication, "drug_name", ""),
                    getattr(medication, "dosage", ""),
                    getattr(medication, "indication", ""),
                    getattr(medication, "direction", ""),
                    getattr(medication, "refill_due", None),
                    getattr(medication, "indication_read", False),
                    getattr(medication, "direction_read", False),
                    getattr(medication, "prescriber", ""),
                    getattr(medication, "notes", ""),
                ]
            )

    workbook.save(path)
    return path
